"""Extract matched general and regular-employee MHLW annual job tables."""

from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

import openpyxl

YEARS = list(range(2011, 2026))
GENERAL_URL = "https://www.mhlw.go.jp/toukei/list/xls/114-1d-01.xlsx"
REGULAR_URL = "https://www.mhlw.go.jp/toukei/list/xls/114-1d-02.xlsx"
METRICS = [
    ("active", "有効求人数"),
    ("new", "新規求人数"),
    ("placed", "就職件数"),
]
PREFECTURES = [
    ("01", "北海道", "北海道"),
    ("02", "青森", "東北"),
    ("03", "岩手", "東北"),
    ("04", "宮城", "東北"),
    ("05", "秋田", "東北"),
    ("06", "山形", "東北"),
    ("07", "福島", "東北"),
    ("08", "茨城", "関東"),
    ("09", "栃木", "関東"),
    ("10", "群馬", "関東"),
    ("11", "埼玉", "関東"),
    ("12", "千葉", "関東"),
    ("13", "東京", "関東"),
    ("14", "神奈川", "関東"),
    ("15", "新潟", "北陸甲信越"),
    ("16", "富山", "北陸甲信越"),
    ("17", "石川", "北陸甲信越"),
    ("18", "福井", "北陸甲信越"),
    ("19", "山梨", "北陸甲信越"),
    ("20", "長野", "北陸甲信越"),
    ("21", "岐阜", "東海"),
    ("22", "静岡", "東海"),
    ("23", "愛知", "東海"),
    ("24", "三重", "東海"),
    ("25", "滋賀", "近畿"),
    ("26", "京都", "近畿"),
    ("27", "大阪", "近畿"),
    ("28", "兵庫", "近畿"),
    ("29", "奈良", "近畿"),
    ("30", "和歌山", "近畿"),
    ("31", "鳥取", "中国"),
    ("32", "島根", "中国"),
    ("33", "岡山", "中国"),
    ("34", "広島", "中国"),
    ("35", "山口", "中国"),
    ("36", "徳島", "四国"),
    ("37", "香川", "四国"),
    ("38", "愛媛", "四国"),
    ("39", "高知", "四国"),
    ("40", "福岡", "九州・沖縄"),
    ("41", "佐賀", "九州・沖縄"),
    ("42", "長崎", "九州・沖縄"),
    ("43", "熊本", "九州・沖縄"),
    ("44", "大分", "九州・沖縄"),
    ("45", "宮崎", "九州・沖縄"),
    ("46", "鹿児島", "九州・沖縄"),
    ("47", "沖縄", "九州・沖縄"),
]


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def read_table(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, list[int]]:
    if sheet.max_row != 49 or sheet.max_column != 16:
        raise ValueError(f"unexpected dimensions for {sheet.title}: {sheet.max_row}x{sheet.max_column}")
    years = [int(str(sheet.cell(1, column).value).replace("年度", "")) for column in range(2, 17)]
    if years != YEARS:
        raise ValueError(f"unexpected years for {sheet.title}: {years}")
    table: dict[str, list[int]] = {}
    for row in range(2, 50):
        label = str(sheet.cell(row, 1).value)
        values = [sheet.cell(row, column).value for column in range(2, 17)]
        if any(not isinstance(value, (int, float)) or value <= 0 for value in values):
            raise ValueError(f"invalid value in {sheet.title}: {label}")
        table[label] = [int(value) for value in values]
    return table


def main() -> None:
    if len(sys.argv) != 4:
        raise SystemExit("usage: extract-source.py GENERAL.xlsx REGULAR.xlsx OUTPUT_DIRECTORY")
    general_path = Path(sys.argv[1])
    regular_path = Path(sys.argv[2])
    output_directory = Path(sys.argv[3])
    general_book = openpyxl.load_workbook(general_path, read_only=True, data_only=True)
    regular_book = openpyxl.load_workbook(regular_path, read_only=True, data_only=True)
    if len(general_book.worksheets) != 3 or len(regular_book.worksheets) != 3:
        raise ValueError("expected exactly three metric sheets per workbook")

    places = [{"id": "00", "name": "全国", "region": "全国"}] + [
        {"id": item_id, "name": name, "region": region}
        for item_id, name, region in PREFECTURES
    ]
    source_labels = ["全国計"] + [f"{name}労働局" for _item_id, name, _region in PREFECTURES]
    records: list[dict[str, object]] = []
    checked = 0

    for metric_index, (metric_id, _metric_name) in enumerate(METRICS):
        general = read_table(general_book.worksheets[metric_index])
        regular = read_table(regular_book.worksheets[metric_index])
        if list(general) != source_labels or list(regular) != source_labels:
            raise ValueError(f"location mismatch in {metric_id}")
        for year_index, year_value in enumerate(YEARS):
            if general["全国計"][year_index] != sum(general[label][year_index] for label in source_labels[1:]):
                raise ValueError(f"general national total mismatch: {metric_id} {year_value}")
            if regular["全国計"][year_index] != sum(regular[label][year_index] for label in source_labels[1:]):
                raise ValueError(f"regular national total mismatch: {metric_id} {year_value}")
        for place, label in zip(places, source_labels, strict=True):
            general_values = general[label]
            regular_values = regular[label]
            if any(regular_value > general_value for regular_value, general_value in zip(regular_values, general_values, strict=True)):
                raise ValueError(f"regular count exceeds general count: {metric_id} {label}")
            checked += len(YEARS)
            records.append({"p": place["id"], "m": metric_id, "g": general_values, "r": regular_values})

    expected_checks = len(places) * len(METRICS) * len(YEARS)
    if checked != expected_checks or len(records) != len(places) * len(METRICS):
        raise ValueError(f"unexpected output size: {checked=} {len(records)=}")
    index = {
        "schemaVersion": 1,
        "asOf": "2026-08-02",
        "edition": "2025年度（令和7年度）まで",
        "years": YEARS,
        "placeCount": len(places),
        "prefectureCount": 47,
        "metricCount": len(METRICS),
        "recordCount": expected_checks,
        "seriesCount": len(records),
        "places": places,
        "metrics": [{"id": item_id, "name": name} for item_id, name in METRICS],
        "sources": [
            {"kind": "general", "url": GENERAL_URL, "sha256": sha256(general_path)},
            {"kind": "regular", "url": REGULAR_URL, "sha256": sha256(regular_path)},
        ],
    }
    output_directory.mkdir(parents=True, exist_ok=True)
    (output_directory / "index.json").write_text(
        json.dumps(index, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    (output_directory / "jobs.json").write_text(
        json.dumps(records, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    print(json.dumps({
        "checked": checked,
        "generalSha256": index["sources"][0]["sha256"],
        "places": len(places),
        "regularSha256": index["sources"][1]["sha256"],
        "series": len(records),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
